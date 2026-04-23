"""스크래핑 결과를 Excel 파일로 저장"""
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_ROW_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def write_to_excel(all_data: dict, output_path: str) -> str:
    """
    all_data: { sheet_name: [row_dict, ...], ... }
    output_path: 저장할 .xlsx 경로
    """
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, rows in all_data.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel 시트 이름 31자 제한

        if not rows:
            ws.append(["데이터 없음"])
            continue

        headers = list(rows[0].keys())
        _write_header(ws, headers)

        for i, row in enumerate(rows, start=2):
            for col_idx, key in enumerate(headers, start=1):
                cell = ws.cell(row=i, column=col_idx, value=str(row.get(key, "")))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = THIN_BORDER
                if i % 2 == 0:
                    cell.fill = ALT_ROW_FILL

        _auto_column_width(ws)

    wb.save(output_path)
    logger.info(f"Excel 저장 완료: {output_path}")
    return output_path


def _write_header(ws, headers: list):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def _auto_column_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                # 멀티라인 시 첫 줄만 기준
                first_line = val.split("\n")[0]
                max_len = max(max_len, len(first_line))
            except Exception:
                pass
        # 최소 10, 최대 60 (한글 고려해 *1.5)
        adjusted = min(max(max_len * 1.5, 10), 60)
        ws.column_dimensions[col_letter].width = adjusted
